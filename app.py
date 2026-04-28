from flask import Flask, render_template, request, redirect, url_for, flash, g, jsonify, Response, send_from_directory
import sqlite3, csv, io, json, os, shutil, threading, time, re, subprocess
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
from db import get_db, init_db, DB_PATH

app = Flask(__name__)
app.secret_key = 'moil-primary-it-tracker-2026'
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB max upload

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'invoices')
IMPORT_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'imports')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'docx', 'doc', 'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


DEVICE_FIELD_ALIASES = {
    'asset_tag':       ['asset tag', 'asset_tag', 'assettag', 'asset', 'tag', 'asset no', 'asset number'],
    'serial_number':   ['serial number', 'serial_number', 'serialnumber', 'serial', 's/n', 'sn', 'serial no'],
    'device_type':     ['device type', 'device_type', 'type', 'category', 'device category'],
    'make':            ['make', 'manufacturer', 'brand', 'mfr'],
    'model':           ['model', 'model number', 'model_number', 'model name'],
    'assigned_to':     ['assigned to', 'assigned_to', 'user', 'owner', 'assignee', 'student', 'staff', 'name'],
    'location':        ['location', 'room', 'area', 'site', 'place'],
    'status':          ['status', 'state', 'device status'],
    'condition':       ['condition', 'device condition', 'physical condition'],
    'os_version':      ['os', 'operating system', 'os version', 'os_version', 'platform'],
    'hostname':        ['hostname', 'host name', 'computer name', 'computername', 'machine name'],
    'purchase_date':   ['purchase date', 'purchase_date', 'bought', 'date purchased', 'date bought', 'buy date'],
    'purchase_price':  ['purchase price', 'purchase_price', 'price', 'cost', 'amount', 'value'],
    'warranty_expiry': ['warranty expiry', 'warranty_expiry', 'warranty', 'warranty end', 'warranty expires', 'warranty date'],
    'funding_source':  ['funding', 'funding source', 'funding_source', 'funded by'],
    'supplier':        ['supplier', 'vendor', 'purchased from', 'retailer'],
    'po_number':       ['po number', 'po_number', 'po', 'purchase order', 'order number'],
    'notes':           ['notes', 'note', 'comments', 'remarks', 'comment'],
}
DEVICE_IMPORT_FIELDS = list(DEVICE_FIELD_ALIASES.keys())


def _auto_map_columns(headers):
    mapping = {}
    for header in headers:
        h = header.lower().strip()
        for field, aliases in DEVICE_FIELD_ALIASES.items():
            if h in aliases:
                mapping[header] = field
                break
        else:
            mapping[header] = ''
    return mapping


def _parse_import_file(filepath):
    ext = filepath.rsplit('.', 1)[-1].lower()
    try:
        if ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        elif ext == 'csv':
            import csv as _csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                all_rows = list(_csv.reader(f))
        else:
            return [], [], {}, ['Unsupported format. Upload .xlsx or .csv']
    except Exception as e:
        return [], [], {}, [f'Could not read file: {e}']

    if not all_rows:
        return [], [], {}, ['File is empty']

    headers = [str(h).strip() if h is not None else f'Col{i+1}' for i, h in enumerate(all_rows[0])]
    data_rows = [[str(c).strip() if c is not None else '' for c in row] for row in all_rows[1:]]
    mapping = _auto_map_columns(headers)
    return data_rows, headers, mapping, []


def _extract_invoice_from_pdf(filepath):
    try:
        result = subprocess.run(['pdftotext', filepath, '-'],
                                capture_output=True, text=True, timeout=30)
        text = result.stdout
    except Exception:
        return {}

    data = {}

    for pattern in [
        r'Total\s+(?:inc(?:luding)?\s+GST)?\s*\$?\s*([\d,]+\.\d{2})',
        r'(?:Grand\s+)?Total\s*:\s*\$?\s*([\d,]+\.\d{2})',
        r'(?:Invoice\s+)?Amount\s*(?:Due)?\s*:\s*\$?\s*([\d,]+\.\d{2})',
        r'TOTAL\s+\$?\s*([\d,]+\.\d{2})',
        r'\$\s*([\d,]+\.\d{2})\s*(?:inc\s+GST|AUD)',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            data['amount'] = m.group(1).replace(',', '')
            break

    for pattern in [
        r'GST\s*(?:10%)?\s*:?\s*\$?\s*([\d,]+\.\d{2})',
        r'G\.?S\.?T\.?\s*:?\s*\$?\s*([\d,]+\.\d{2})',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            data['gst'] = m.group(1).replace(',', '')
            break

    date_tries = [
        (r'\b(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})\b', '%d %B %Y'),
        (r'\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})\b', '%d %b %Y'),
        (r'\b(\d{4})-(\d{2})-(\d{2})\b', '%Y-%m-%d'),
    ]
    for pattern, fmt in date_tries:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            try:
                d = datetime.strptime(m.group(0), fmt)
                data['date'] = d.strftime('%Y-%m-%d')
                break
            except Exception:
                pass
    if 'date' not in data:
        m = re.search(r'\b(\d{1,2})/(\d{1,2})/(\d{4})\b', text)
        if m:
            try:
                d = datetime.strptime(m.group(0), '%d/%m/%Y')
                data['date'] = d.strftime('%Y-%m-%d')
            except Exception:
                pass

    for pattern in [
        r'Invoice\s+(?:No\.?|Number|#)\s*:?\s*([A-Z0-9][-A-Z0-9/]{1,30})',
        r'INV[-#]?\s*([A-Z0-9][-A-Z0-9]{1,20})',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            data['invoice_number'] = m.group(1).strip()
            break

    for pattern in [
        r'(?:PO|Purchase\s+Order)\s+(?:No\.?|Number|#)?\s*:?\s*([A-Z0-9][-A-Z0-9]{3,20})',
        r'P\.?O\.?\s+(?:No\.?|#)?\s*:?\s*([A-Z0-9][-A-Z0-9]{3,20})',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            data['po_number'] = m.group(1).strip()
            break

    for pattern in [
        r'(?:From|Supplier|Vendor|Sold\s+By|Bill\s+To)\s*:?\s*([^\n]{4,80})',
        r'^([A-Z][A-Za-z\s&.,()-]{3,60}(?:Pty\.?\s*Ltd\.?|Solutions|Services|Technologies|Digital|Australia))',
    ]:
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if m:
            v = m.group(1).strip()
            if len(v) > 3:
                data['vendor'] = v[:100]
                break

    for pattern in [
        r'(?:Description|Item|Product|For|Re)\s*:?\s*([^\n]{5,200})',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            desc = m.group(1).strip()
            if len(desc) > 4:
                data['description'] = desc[:200]
                break

    tl = text.lower()
    if any(w in tl for w in ['laptop', 'computer', 'pc', 'monitor', 'ipad', 'tablet', 'server', 'printer', 'chromebook']):
        data['category'] = 'ICT Equipment'
    elif any(w in tl for w in ['software', 'licence', 'license', 'subscription', 'microsoft 365', 'adobe']):
        data['category'] = 'Software / Licences'
    elif any(w in tl for w in ['repair', 'maintenance', 'service call', 'technical support']):
        data['category'] = 'Repairs & Maintenance'
    elif any(w in tl for w in ['cable', 'mouse', 'keyboard', 'headset', 'peripheral', 'case', 'hdmi', 'adapter', 'charger']):
        data['category'] = 'Accessories & Peripherals'
    elif any(w in tl for w in ['switch', 'router', 'network', 'wifi', 'ethernet', 'access point', 'firewall']):
        data['category'] = 'Networking'
    elif any(w in tl for w in ['desk', 'chair', 'table', 'furniture']):
        data['category'] = 'Furniture'
    else:
        data['category'] = 'ICT Equipment'

    return data


def _import_clean_float(v):
    try:
        return float(str(v).replace('$', '').replace(',', '').strip())
    except Exception:
        return None


def _import_clean_date(v):
    if not v:
        return None
    for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y', '%d %B %Y'):
        try:
            return datetime.strptime(str(v).strip(), fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return None


def _import_clean_status(v):
    v = str(v).lower().strip()
    return v if v in STATUSES else 'available'


def _import_clean_condition(v):
    for c in CONDITIONS:
        if c.lower() == str(v).lower().strip():
            return c
    return 'Good'

DEVICE_TYPES = ['Student Laptop','Staff Laptop','Teacher Laptop','Desktop PC',
                'New PC Lab PC','Monitor','iPad','Other']
MAKES = ['Dell','Lenovo','Apple','HP','Acer','Samsung','Philips','Microsoft','Other']
STATUSES = ['assigned','available','maintenance','retired']
CONDITIONS = ['New','Excellent','Good','Fair','Poor']
OS_VERSIONS = ['Windows 11','Windows 10','macOS','iOS','iPadOS','Chrome OS','Linux','N/A']
FUNDING = ['NT Government','School Budget','Grant Funded','Donated','Unknown']
LOCATIONS = ['Staff Room','Library','Front Office',"Principal's Office",
             'Business Manager Room','ICT Room','Corner Room','Desert Rose',
             'Student Support','Server Room','Storage',
             'Classroom 1','Classroom 2','Classroom 3','Classroom 4',
             'Classroom 5','Classroom 6','Other']
MAINTENANCE_TYPES = ['Repair','Reimage','Hardware Upgrade','Software Install','Inspection','Other']
LICENCE_TYPES = ['Per Device','Per User','Per Mac','Site Licence','Subscription','One-Time']
VENDORS_LIST = ['Microsoft','Adobe','Apple','Dell','Lenovo','NTG ICT','Other']
ACCESSORY_TYPES = ['HDMI Cable','USB-C Cable','USB-A Cable','DisplayPort Cable',
                   'Power Cable','Charging Adapter','USB-C Charger','Lightning Cable',
                   'Docking Station','USB Hub','Keyboard','Mouse','Webcam',
                   'Headset','Laptop Bag','Tablet Case','Monitor Stand','Other']

INVOICE_CATEGORIES = ['ICT Equipment','Software / Licences','Repairs & Maintenance',
                      'Accessories & Peripherals','Networking','Furniture','Other']
PAYMENT_STATUSES = ['paid','pending','overdue','cancelled']
INVOICE_LOCATIONS = ['ICT Room','Staff Room','Library','Front Office',
                     "Principal's Office",'Business Manager Room','Corner Room',
                     'Desert Rose','Student Support','Server Room',
                     'Classroom 1','Classroom 2','Classroom 3','Classroom 4',
                     'Classroom 5','Classroom 6','Classrooms (General)','Storage','Other']
TROLLEYS = ['Trolley 1 (Green)',
            'Trolley 2 (Grey)','Trolley 3 (Orange)',
            'Trolley 4 (White)']
FURNITURE_CATEGORIES = ['Desk','Chair','Table','Cabinet','Shelving','Whiteboard',
                        'Bookcase','Storage','Display','Other']
FURNITURE_STATUSES = ['active','in storage','disposed','on loan']

WARRANTY_WARN_DAYS = 90
LICENCE_WARN_DAYS = 60
DEVICE_AGE_WARN_YEARS = 5


@app.before_request
def open_db():
    g.db = get_db()


@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()


def today():
    return date.today().isoformat()


def days_until(date_str):
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str[:10], '%Y-%m-%d').date()
        return (d - date.today()).days
    except:
        return None


def is_expired(date_str):
    d = days_until(date_str)
    return d is not None and d < 0


def is_expiring_soon(date_str, warn_days):
    d = days_until(date_str)
    return d is not None and 0 <= d <= warn_days


def fmt_date(date_str):
    if not date_str:
        return '—'
    try:
        d = datetime.strptime(date_str[:10], '%Y-%m-%d')
        return d.strftime('%d %b %Y')
    except:
        return date_str


def fmt_currency(val):
    if val is None:
        return '—'
    try:
        return f'${float(val):,.2f}'
    except:
        return str(val)


def display_model(make, model):
    """Return make+model string without duplicating make if model already starts with it."""
    make = (make or '').strip()
    model = (model or '').strip()
    if not model:
        return make or '—'
    if make and model.lower().startswith(make.lower()):
        return model
    return f'{make} {model}' if make else model


app.jinja_env.globals.update(
    display_model=display_model,
    fmt_date=fmt_date,
    fmt_currency=fmt_currency,
    days_until=days_until,
    is_expired=is_expired,
    is_expiring_soon=is_expiring_soon,
    today=today,
    WARRANTY_WARN_DAYS=WARRANTY_WARN_DAYS,
    LICENCE_WARN_DAYS=LICENCE_WARN_DAYS,
    DEVICE_TYPES=DEVICE_TYPES,
    MAKES=MAKES,
    STATUSES=STATUSES,
    CONDITIONS=CONDITIONS,
    OS_VERSIONS=OS_VERSIONS,
    FUNDING=FUNDING,
    LOCATIONS=LOCATIONS,
    MAINTENANCE_TYPES=MAINTENANCE_TYPES,
    LICENCE_TYPES=LICENCE_TYPES,
    VENDORS_LIST=VENDORS_LIST,
    ACCESSORY_TYPES=ACCESSORY_TYPES,
    INVOICE_CATEGORIES=INVOICE_CATEGORIES,
    PAYMENT_STATUSES=PAYMENT_STATUSES,
    INVOICE_LOCATIONS=INVOICE_LOCATIONS,
    FURNITURE_CATEGORIES=FURNITURE_CATEGORIES,
    FURNITURE_STATUSES=FURNITURE_STATUSES,
    TROLLEYS=TROLLEYS,
)


# ─── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    db = g.db
    devices = db.execute("SELECT * FROM devices").fetchall()
    licences = db.execute("SELECT * FROM licences").fetchall()
    maintenance = db.execute(
        "SELECT m.*, d.asset_tag, d.serial_number, d.device_type FROM maintenance m "
        "LEFT JOIN devices d ON m.device_id=d.id ORDER BY m.date DESC LIMIT 8"
    ).fetchall()

    total = len(devices)
    assigned = sum(1 for d in devices if d['status'] == 'assigned')
    available = sum(1 for d in devices if d['status'] == 'available')
    in_maint = sum(1 for d in devices if d['status'] == 'maintenance')
    retired = sum(1 for d in devices if d['status'] == 'retired')

    warranty_expiring = [d for d in devices if is_expiring_soon(d['warranty_expiry'], WARRANTY_WARN_DAYS)]
    warranty_expired = [d for d in devices if is_expired(d['warranty_expiry'])]
    lic_expiring = [l for l in licences if is_expiring_soon(l['renewal_date'], LICENCE_WARN_DAYS)]
    lic_expired = [l for l in licences if is_expired(l['renewal_date'])]

    type_counts = {}
    for t in DEVICE_TYPES:
        type_counts[t] = sum(1 for d in devices if d['device_type'] == t)

    return render_template('dashboard.html',
        total=total, assigned=assigned, available=available,
        in_maint=in_maint, retired=retired,
        warranty_expiring=warranty_expiring, warranty_expired=warranty_expired,
        lic_expiring=lic_expiring, lic_expired=lic_expired,
        type_counts=type_counts,
        recent_maintenance=maintenance,
        licences=licences,
    )


# ─── Devices ──────────────────────────────────────────────────────────────────

@app.route('/devices')
def devices():
    db = g.db
    search = request.args.get('q', '').strip()
    dtype = request.args.get('type', '')
    status = request.args.get('status', '')
    condition = request.args.get('condition', '')
    trolley = request.args.get('trolley', '')

    query = "SELECT * FROM devices WHERE 1=1"
    params = []

    if search:
        query += """ AND (asset_tag LIKE ? OR serial_number LIKE ? OR hostname LIKE ?
                     OR assigned_to LIKE ? OR make LIKE ? OR model LIKE ?)"""
        p = f'%{search}%'
        params.extend([p, p, p, p, p, p])
    if dtype:
        query += " AND device_type=?"
        params.append(dtype)
    if status:
        query += " AND status=?"
        params.append(status)
    if condition:
        query += " AND condition=?"
        params.append(condition)
    if trolley:
        query += " AND trolley=?"
        params.append(trolley)

    query += " ORDER BY device_type, asset_tag"
    rows = db.execute(query, params).fetchall()

    return render_template('devices.html', devices=rows,
        search=search, dtype=dtype, status=status, condition=condition, trolley=trolley)


@app.route('/devices/new', methods=['GET', 'POST'])
def device_new():
    if request.method == 'POST':
        _save_device(None)
        flash('Device added successfully.', 'success')
        return redirect(url_for('devices'))
    return render_template('device_form.html', device=None, title='Add Device')


@app.route('/devices/<int:id>')
def device_detail(id):
    db = g.db
    device = db.execute("SELECT * FROM devices WHERE id=?", (id,)).fetchone()
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('devices'))
    maint = db.execute(
        "SELECT * FROM maintenance WHERE device_id=? ORDER BY date DESC", (id,)
    ).fetchall()
    checkouts = db.execute(
        "SELECT * FROM checkout WHERE device_id=? ORDER BY checked_out_at DESC LIMIT 10", (id,)
    ).fetchall()
    accessories = db.execute(
        "SELECT * FROM accessories WHERE device_id=?", (id,)
    ).fetchall()
    return render_template('device_detail.html', device=device,
                           maintenance=maint, checkouts=checkouts, accessories=accessories)


@app.route('/devices/<int:id>/edit', methods=['GET', 'POST'])
def device_edit(id):
    db = g.db
    device = db.execute("SELECT * FROM devices WHERE id=?", (id,)).fetchone()
    if not device:
        return redirect(url_for('devices'))
    if request.method == 'POST':
        _save_device(id)
        flash('Device updated.', 'success')
        return redirect(url_for('device_detail', id=id))
    return render_template('device_form.html', device=device, title='Edit Device')


@app.route('/devices/<int:id>/delete', methods=['POST'])
def device_delete(id):
    g.db.execute("DELETE FROM devices WHERE id=?", (id,))
    g.db.commit()
    flash('Device deleted.', 'info')
    return redirect(url_for('devices'))


@app.route('/devices/import', methods=['GET', 'POST'])
def device_import():
    if request.method == 'GET':
        return render_template('device_import.html', step='upload')

    action = request.form.get('action', 'preview')

    if action == 'preview':
        if 'import_file' not in request.files:
            flash('No file selected.', 'error')
            return render_template('device_import.html', step='upload')
        f = request.files['import_file']
        if not f or not f.filename:
            flash('No file selected.', 'error')
            return render_template('device_import.html', step='upload')
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ('xlsx', 'xls', 'csv'):
            flash('Please upload an Excel (.xlsx) or CSV (.csv) file.', 'error')
            return render_template('device_import.html', step='upload')

        os.makedirs(IMPORT_FOLDER, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{ts}_{secure_filename(f.filename)}"
        filepath = os.path.join(IMPORT_FOLDER, filename)
        f.save(filepath)

        rows, headers, mapping, errors = _parse_import_file(filepath)
        if errors:
            flash(errors[0], 'error')
            return render_template('device_import.html', step='upload')

        return render_template('device_import.html',
            step='preview',
            headers=headers,
            preview=rows[:10],
            mapping=mapping,
            total_rows=len(rows),
            import_filename=filename,
            device_import_fields=DEVICE_IMPORT_FIELDS,
        )

    elif action == 'import':
        filename = request.form.get('import_filename', '').strip()
        if not filename or '/' in filename or '..' in filename:
            flash('Invalid import file reference.', 'error')
            return redirect(url_for('device_import'))

        filepath = os.path.join(IMPORT_FOLDER, filename)
        if not os.path.exists(filepath):
            flash('Import file not found — please re-upload.', 'error')
            return redirect(url_for('device_import'))

        rows, headers, _, errors = _parse_import_file(filepath)
        if errors:
            flash(errors[0], 'error')
            return redirect(url_for('device_import'))

        col_map = {}
        for i in range(len(headers)):
            field = request.form.get(f'map_{i}', '').strip()
            if field:
                col_map[i] = field

        db = g.db
        imported = 0
        skipped = 0
        now = datetime.now().isoformat()

        for row in rows:
            device = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    val = row[col_idx].strip()
                    if val:
                        device[field_name] = val

            if not device:
                skipped += 1
                continue

            asset_tag = device.get('asset_tag', '').strip()
            serial_number = device.get('serial_number', '').strip()

            if asset_tag and db.execute("SELECT 1 FROM devices WHERE asset_tag=?", (asset_tag,)).fetchone():
                skipped += 1
                continue
            if serial_number and not asset_tag and db.execute("SELECT 1 FROM devices WHERE serial_number=?", (serial_number,)).fetchone():
                skipped += 1
                continue

            db.execute("""INSERT INTO devices
                (asset_tag,serial_number,device_type,make,model,assigned_to,location,
                 status,condition,os_version,hostname,purchase_date,purchase_price,
                 warranty_expiry,funding_source,supplier,po_number,notes,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                asset_tag or None,
                serial_number or None,
                device.get('device_type', 'Other').strip() or 'Other',
                device.get('make', '').strip(),
                device.get('model', '').strip(),
                device.get('assigned_to', '').strip(),
                device.get('location', '').strip(),
                _import_clean_status(device.get('status', 'available')),
                _import_clean_condition(device.get('condition', 'Good')),
                device.get('os_version', '').strip(),
                device.get('hostname', '').strip(),
                _import_clean_date(device.get('purchase_date', '')),
                _import_clean_float(device.get('purchase_price', '')),
                _import_clean_date(device.get('warranty_expiry', '')),
                device.get('funding_source', '').strip(),
                device.get('supplier', '').strip(),
                device.get('po_number', '').strip(),
                device.get('notes', '').strip(),
                now,
            ))
            imported += 1

        db.commit()
        flash(f'Import complete: {imported} device{"s" if imported != 1 else ""} added, {skipped} skipped (duplicates or empty rows).', 'success')
        return redirect(url_for('devices'))

    return redirect(url_for('device_import'))


@app.route('/devices/import/template')
def device_import_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Device Import'
    headers = ['Asset Tag', 'Serial Number', 'Device Type', 'Make', 'Model',
               'Assigned To', 'Location', 'Status', 'Condition', 'OS Version',
               'Hostname', 'Purchase Date', 'Purchase Price', 'Warranty Expiry',
               'Funding Source', 'Supplier', 'PO Number', 'Notes']
    example = ['MPS-001', 'SN12345678', 'Student Laptop', 'Dell', 'Latitude 3140',
               'Student Name', 'ICT Room', 'assigned', 'Good', 'Windows 11',
               'LAPTOP-01', '15/01/2024', '1200.00', '15/01/2027',
               'NT Government', 'Territory Technology Solutions', 'PO15500143', '']
    ws.append(headers)
    ws.append(example)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=device_import_template.xlsx'})


def _save_device(id):
    f = request.form
    db = g.db
    vals = (
        f.get('asset_tag','').strip() or None,
        f.get('serial_number','').strip() or None,
        f.get('device_type',''), f.get('make',''), f.get('model','').strip(),
        f.get('assigned_to','').strip(), f.get('location',''),
        f.get('status','available'), f.get('condition','Good'),
        f.get('os_version',''), f.get('storage',''),
        f.get('purchase_date') or None, f.get('purchase_price') or None,
        f.get('warranty_expiry') or None, f.get('funding_source',''),
        f.get('supplier','').strip() or None,
        f.get('po_number','').strip() or None,
        f.get('invoice_number','').strip() or None,
        f.get('trolley','').strip() or None,
        f.get('hostname','').strip() or None,
        1 if f.get('domain_joined') else 0,
        1 if f.get('bitlocker_enabled') else 0,
        1 if f.get('mdm_enrolled') else 0,
        f.get('last_reimaged') or None,
        f.get('charger_type',''), 1 if f.get('charger_included') else 0,
        1 if f.get('case_loan') else 0,
        f.get('notes','').strip() or None,
        datetime.now().isoformat(),
    )
    if id is None:
        db.execute("""INSERT INTO devices
            (asset_tag,serial_number,device_type,make,model,assigned_to,location,
             status,condition,os_version,storage,purchase_date,purchase_price,
             warranty_expiry,funding_source,supplier,po_number,invoice_number,trolley,hostname,
             domain_joined,bitlocker_enabled,mdm_enrolled,last_reimaged,
             charger_type,charger_included,case_loan,notes,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", vals)
    else:
        db.execute("""UPDATE devices SET
            asset_tag=?,serial_number=?,device_type=?,make=?,model=?,
            assigned_to=?,location=?,status=?,condition=?,os_version=?,storage=?,
            purchase_date=?,purchase_price=?,warranty_expiry=?,funding_source=?,
            supplier=?,po_number=?,invoice_number=?,trolley=?,hostname=?,domain_joined=?,bitlocker_enabled=?,
            mdm_enrolled=?,last_reimaged=?,charger_type=?,charger_included=?,
            case_loan=?,notes=?,updated_at=? WHERE id=?""", vals + (id,))
    db.commit()


# ─── Licences ─────────────────────────────────────────────────────────────────

@app.route('/licences')
def licences():
    rows = g.db.execute("SELECT * FROM licences ORDER BY software").fetchall()
    return render_template('licences.html', licences=rows)


@app.route('/licences/new', methods=['GET', 'POST'])
def licence_new():
    if request.method == 'POST':
        _save_licence(None)
        flash('Licence added.', 'success')
        return redirect(url_for('licences'))
    return render_template('licence_form.html', licence=None, title='Add Licence')


@app.route('/licences/<int:id>/edit', methods=['GET', 'POST'])
def licence_edit(id):
    db = g.db
    lic = db.execute("SELECT * FROM licences WHERE id=?", (id,)).fetchone()
    if request.method == 'POST':
        _save_licence(id)
        flash('Licence updated.', 'success')
        return redirect(url_for('licences'))
    return render_template('licence_form.html', licence=lic, title='Edit Licence')


@app.route('/licences/<int:id>/delete', methods=['POST'])
def licence_delete(id):
    g.db.execute("DELETE FROM licences WHERE id=?", (id,))
    g.db.commit()
    flash('Licence deleted.', 'info')
    return redirect(url_for('licences'))


def _save_licence(id):
    f = request.form
    db = g.db
    vals = (
        f.get('software','').strip(), f.get('vendor',''),
        f.get('licence_type',''), f.get('seats') or None,
        f.get('cost_per_unit') or None, f.get('total_cost') or None,
        f.get('billing_year') or None, f.get('renewal_date') or None,
        f.get('status','active'), f.get('assigned_to','').strip(),
        f.get('notes','').strip(), datetime.now().isoformat(),
    )
    if id is None:
        db.execute("""INSERT INTO licences
            (software,vendor,licence_type,seats,cost_per_unit,total_cost,
             billing_year,renewal_date,status,assigned_to,notes,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", vals)
    else:
        db.execute("""UPDATE licences SET
            software=?,vendor=?,licence_type=?,seats=?,cost_per_unit=?,total_cost=?,
            billing_year=?,renewal_date=?,status=?,assigned_to=?,notes=?,updated_at=?
            WHERE id=?""", vals + (id,))
    db.commit()


# ─── Maintenance ──────────────────────────────────────────────────────────────

@app.route('/maintenance')
def maintenance():
    rows = g.db.execute("""
        SELECT m.*, d.asset_tag, d.serial_number, d.device_type, d.model
        FROM maintenance m LEFT JOIN devices d ON m.device_id=d.id
        ORDER BY m.date DESC
    """).fetchall()
    devices = g.db.execute("SELECT id, asset_tag, serial_number, device_type FROM devices ORDER BY device_type, asset_tag").fetchall()
    return render_template('maintenance.html', records=rows, devices=devices)


@app.route('/maintenance/new', methods=['GET', 'POST'])
def maintenance_new():
    if request.method == 'POST':
        f = request.form
        db = g.db
        db.execute("""INSERT INTO maintenance (device_id,date,type,technician,cost,status_after,notes)
            VALUES (?,?,?,?,?,?,?)""", (
            f.get('device_id') or None, f.get('date') or today(),
            f.get('type',''), f.get('technician','').strip(),
            f.get('cost') or None, f.get('status_after',''),
            f.get('notes','').strip()
        ))
        if f.get('status_after') and f.get('device_id'):
            db.execute("UPDATE devices SET status=?,updated_at=? WHERE id=?",
                       (f['status_after'], datetime.now().isoformat(), f['device_id']))
        db.commit()
        flash('Maintenance record added.', 'success')
        return redirect(url_for('maintenance'))
    devices = g.db.execute("SELECT id, asset_tag, serial_number, device_type FROM devices ORDER BY device_type, asset_tag").fetchall()
    pre_device = request.args.get('device_id')
    return render_template('maintenance_form.html', devices=devices, pre_device=pre_device)


@app.route('/maintenance/<int:id>/delete', methods=['POST'])
def maintenance_delete(id):
    g.db.execute("DELETE FROM maintenance WHERE id=?", (id,))
    g.db.commit()
    flash('Record deleted.', 'info')
    return redirect(url_for('maintenance'))


# ─── Checkout ─────────────────────────────────────────────────────────────────

@app.route('/checkout')
def checkout():
    active = g.db.execute("""
        SELECT c.*, d.asset_tag, d.serial_number, d.device_type FROM checkout c
        LEFT JOIN devices d ON c.device_id=d.id WHERE c.status='out' ORDER BY c.checked_out_at DESC
    """).fetchall()
    history = g.db.execute("""
        SELECT c.*, d.asset_tag, d.serial_number, d.device_type FROM checkout c
        LEFT JOIN devices d ON c.device_id=d.id WHERE c.status='returned'
        ORDER BY c.returned_at DESC LIMIT 50
    """).fetchall()
    devices = g.db.execute(
        "SELECT id, asset_tag, serial_number, device_type FROM devices ORDER BY device_type, asset_tag"
    ).fetchall()
    return render_template('checkout.html', active=active, history=history, devices=devices)


@app.route('/checkout/new', methods=['POST'])
def checkout_new():
    f = request.form
    db = g.db
    db.execute("""INSERT INTO checkout (device_id,borrower_name,borrower_class,
        checked_out_at,expected_return,notes,status)
        VALUES (?,?,?,?,?,?,'out')""", (
        f.get('device_id') or None, f.get('borrower_name','').strip(),
        f.get('borrower_class','').strip(), f.get('checked_out_at') or today(),
        f.get('expected_return') or None, f.get('notes','').strip()
    ))
    if f.get('device_id'):
        db.execute("UPDATE devices SET status='assigned', assigned_to=?, updated_at=? WHERE id=?",
                   (f.get('borrower_name',''), datetime.now().isoformat(), f['device_id']))
    db.commit()
    flash(f"Device checked out to {f.get('borrower_name','')}", 'success')
    return redirect(url_for('checkout'))


@app.route('/checkout/<int:id>/return', methods=['POST'])
def checkout_return(id):
    db = g.db
    row = db.execute("SELECT * FROM checkout WHERE id=?", (id,)).fetchone()
    if row:
        db.execute("UPDATE checkout SET status='returned', returned_at=? WHERE id=?",
                   (datetime.now().isoformat(), id))
        if row['device_id']:
            condition = request.form.get('condition', 'Good')
            db.execute("UPDATE devices SET status='available', assigned_to='', condition=?, updated_at=? WHERE id=?",
                       (condition, datetime.now().isoformat(), row['device_id']))
        db.commit()
        flash('Device returned.', 'success')
    return redirect(url_for('checkout'))


# ─── Vendors ──────────────────────────────────────────────────────────────────

@app.route('/vendors')
def vendors():
    rows = g.db.execute("SELECT * FROM vendors ORDER BY name").fetchall()
    return render_template('vendors.html', vendors=rows)


@app.route('/vendors/new', methods=['GET', 'POST'])
def vendor_new():
    if request.method == 'POST':
        _save_vendor(None)
        flash('Vendor added.', 'success')
        return redirect(url_for('vendors'))
    return render_template('vendor_form.html', vendor=None, title='Add Vendor')


@app.route('/vendors/<int:id>/edit', methods=['GET', 'POST'])
def vendor_edit(id):
    v = g.db.execute("SELECT * FROM vendors WHERE id=?", (id,)).fetchone()
    if request.method == 'POST':
        _save_vendor(id)
        flash('Vendor updated.', 'success')
        return redirect(url_for('vendors'))
    return render_template('vendor_form.html', vendor=v, title='Edit Vendor')


@app.route('/vendors/<int:id>/delete', methods=['POST'])
def vendor_delete(id):
    g.db.execute("DELETE FROM vendors WHERE id=?", (id,))
    g.db.commit()
    flash('Vendor deleted.', 'info')
    return redirect(url_for('vendors'))


def _save_vendor(id):
    f = request.form
    db = g.db
    vals = (f.get('name','').strip(), f.get('contact','').strip(), f.get('phone','').strip(),
            f.get('email','').strip(), f.get('website','').strip(),
            f.get('categories','').strip(), f.get('account_number','').strip(),
            f.get('notes','').strip())
    if id is None:
        db.execute("INSERT INTO vendors (name,contact,phone,email,website,categories,account_number,notes) VALUES (?,?,?,?,?,?,?,?)", vals)
    else:
        db.execute("UPDATE vendors SET name=?,contact=?,phone=?,email=?,website=?,categories=?,account_number=?,notes=? WHERE id=?", vals + (id,))
    db.commit()


# ─── Export ───────────────────────────────────────────────────────────────────

@app.route('/export')
def export():
    db = g.db
    dev_count = db.execute("SELECT COUNT(*) as c FROM devices").fetchone()['c']
    lic_count = db.execute("SELECT COUNT(*) as c FROM licences").fetchone()['c']
    mnt_count = db.execute("SELECT COUNT(*) as c FROM maintenance").fetchone()['c']
    row = db.execute("SELECT value FROM settings WHERE key='last_backup'").fetchone()
    last_backup = row['value'] if row else None
    return render_template('export.html', dev_count=dev_count,
                           lic_count=lic_count, mnt_count=mnt_count,
                           last_backup=last_backup)


@app.route('/export/devices.csv')
def export_devices():
    rows = g.db.execute("SELECT * FROM devices ORDER BY device_type, asset_tag").fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Asset Tag','Serial','Device Type','Make','Model','Assigned To',
                'Location','Status','Condition','OS','Purchase Date','Warranty Expiry',
                'Domain','BitLocker','MDM','Notes'])
    for r in rows:
        w.writerow([r['asset_tag'],r['serial_number'],r['device_type'],r['make'],r['model'],
                    r['assigned_to'],r['location'],r['status'],r['condition'],
                    r['os_version'],r['purchase_date'],r['warranty_expiry'],
                    'Yes' if r['domain_joined'] else 'No',
                    'Yes' if r['bitlocker_enabled'] else 'No',
                    'Yes' if r['mdm_enrolled'] else 'No',r['notes']])
    return Response(out.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=devices.csv'})


@app.route('/export/licences.csv')
def export_licences():
    rows = g.db.execute("SELECT * FROM licences ORDER BY software").fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Software','Vendor','Type','Seats','Cost/Unit','Total','Renewal','Status','Notes'])
    for r in rows:
        w.writerow([r['software'],r['vendor'],r['licence_type'],r['seats'],
                    r['cost_per_unit'],r['total_cost'],r['renewal_date'],r['status'],r['notes']])
    return Response(out.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=licences.csv'})


@app.route('/export/maintenance.csv')
def export_maintenance():
    rows = g.db.execute("""
        SELECT m.*, d.asset_tag, d.device_type FROM maintenance m
        LEFT JOIN devices d ON m.device_id=d.id ORDER BY m.date DESC
    """).fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Date','Asset Tag','Device Type','Type','Technician','Cost','Notes'])
    for r in rows:
        w.writerow([r['date'],r['asset_tag'],r['device_type'],r['type'],
                    r['technician'],r['cost'],r['notes']])
    return Response(out.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=maintenance.csv'})


@app.route('/export/backup/download')
def export_backup_download():
    with open(DB_PATH, 'rb') as f:
        data = f.read()
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(data, mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename=tracker_backup_{ts}.db'})


@app.route('/export/backup/manual', methods=['POST'])
def export_backup_manual():
    _do_backup()
    flash('Backup saved to backups/ folder.', 'success')
    return redirect(url_for('export'))


# ─── Budget / Invoices ────────────────────────────────────────────────────────

@app.route('/budget')
def budget():
    db = g.db
    year = request.args.get('year', date.today().year, type=int)

    if request.args.get('export') == 'csv':
        return _export_invoices_csv(db, year)

    search   = request.args.get('q', '').strip()
    vendor_f = request.args.get('vendor', '')
    cat_f    = request.args.get('cat', '')
    status_f = request.args.get('status', '')
    loc_f    = request.args.get('location', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    # Filtered invoice query
    query  = "SELECT * FROM invoices WHERE strftime('%Y', date)=?"
    params = [str(year)]
    if search:
        p = f'%{search}%'
        query += " AND (description LIKE ? OR vendor LIKE ? OR invoice_number LIKE ? OR po_number LIKE ?)"
        params.extend([p, p, p, p])
    if vendor_f:
        query += " AND vendor=?";     params.append(vendor_f)
    if cat_f:
        query += " AND category=?";   params.append(cat_f)
    if status_f:
        query += " AND payment_status=?"; params.append(status_f)
    if loc_f:
        query += " AND location=?";   params.append(loc_f)
    if date_from:
        query += " AND date >= ?";    params.append(date_from)
    if date_to:
        query += " AND date <= ?";    params.append(date_to)
    query += " ORDER BY date DESC"
    invoices = db.execute(query, params).fetchall()

    # All invoices for the year (unfiltered — for charts)
    all_year = db.execute(
        "SELECT * FROM invoices WHERE strftime('%Y', date)=? ORDER BY date",
        (str(year),)
    ).fetchall()

    years = db.execute(
        "SELECT DISTINCT strftime('%Y', date) as yr FROM invoices ORDER BY yr DESC"
    ).fetchall()
    year_list = [r['yr'] for r in years]
    if str(year) not in year_list:
        year_list.insert(0, str(year))

    # Filtered totals
    total         = sum(r['amount'] for r in invoices)
    paid_total    = sum(r['amount'] for r in invoices if r['payment_status'] == 'paid')
    pending_total = sum(r['amount'] for r in invoices if r['payment_status'] == 'pending')
    overdue_total = sum(r['amount'] for r in invoices if r['payment_status'] == 'overdue')
    by_cat = {}
    for r in invoices:
        by_cat[r['category']] = by_cat.get(r['category'], 0) + r['amount']

    # Budget target for this year
    tgt_row = db.execute("SELECT * FROM budget_targets WHERE year=?", (year,)).fetchone()
    budget_target = tgt_row['target'] if tgt_row else None
    budget_pct    = round(paid_total / budget_target * 100, 1) if budget_target else None

    # Monthly chart (all year, paid only for accurate spend)
    monthly = {}
    for r in all_year:
        if r['date'] and r['payment_status'] == 'paid':
            k = r['date'][:7]
            monthly[k] = monthly.get(k, 0) + r['amount']
    month_labels = [f"{year}-{m:02d}" for m in range(1, 13)]
    month_data   = [round(monthly.get(k, 0), 2) for k in month_labels]

    # Vendor chart (all year)
    by_vendor = {}
    for r in all_year:
        v = (r['vendor'] or 'Unknown').strip()
        by_vendor[v] = by_vendor.get(v, 0) + r['amount']
    top_vendors = sorted(by_vendor.items(), key=lambda x: x[1], reverse=True)[:8]

    # Category chart (all year)
    by_cat_all = {}
    for r in all_year:
        by_cat_all[r['category']] = by_cat_all.get(r['category'], 0) + r['amount']

    # Location chart (all year)
    by_loc = {}
    for r in all_year:
        loc = (r['location'] or 'Unspecified').strip() or 'Unspecified'
        by_loc[loc] = by_loc.get(loc, 0) + r['amount']
    by_loc_sorted = sorted(by_loc.items(), key=lambda x: x[1], reverse=True)

    # Dropdown helpers
    vendor_rows = db.execute(
        "SELECT DISTINCT vendor FROM invoices WHERE vendor != '' AND vendor IS NOT NULL ORDER BY vendor"
    ).fetchall()
    vendors_list = [r['vendor'] for r in vendor_rows]

    loc_rows = db.execute(
        "SELECT DISTINCT location FROM invoices WHERE location != '' AND location IS NOT NULL ORDER BY location"
    ).fetchall()
    locations_used = [r['location'] for r in loc_rows]

    return render_template('budget.html',
        invoices=invoices,
        year=year, year_list=year_list,
        total=total, paid_total=paid_total,
        pending_total=pending_total, overdue_total=overdue_total,
        by_cat=by_cat,
        budget_target=budget_target, budget_pct=budget_pct,
        month_labels=json.dumps(month_labels),
        month_data=json.dumps(month_data),
        top_vendors=json.dumps([[k, round(v, 2)] for k, v in top_vendors]),
        by_cat_all=json.dumps([[k, round(v, 2)] for k, v in by_cat_all.items()]),
        by_loc=json.dumps([[k, round(v, 2)] for k, v in by_loc_sorted]),
        vendors_list=vendors_list, locations_used=locations_used,
        search=search, vendor_f=vendor_f, cat_f=cat_f,
        status_f=status_f, loc_f=loc_f,
        date_from=date_from, date_to=date_to,
    )


@app.route('/budget/target', methods=['POST'])
def budget_target_save():
    year   = request.form.get('year', type=int)
    target = request.form.get('target', type=float)
    notes  = request.form.get('notes', '').strip()
    if year and target is not None:
        g.db.execute(
            "INSERT INTO budget_targets (year, target, notes) VALUES (?,?,?) "
            "ON CONFLICT(year) DO UPDATE SET target=excluded.target, notes=excluded.notes",
            (year, target, notes)
        )
        g.db.commit()
        flash(f'Budget target for {year} set to ${target:,.2f}.', 'success')
    return redirect(url_for('budget', year=year))


@app.route('/budget/extract-pdf', methods=['POST'])
def invoice_extract_pdf():
    if 'invoice_file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['invoice_file']
    if not file or not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'error': 'File type not allowed'}), 400

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{ts}_{secure_filename(file.filename)}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    data = {}
    if ext == 'pdf':
        data = _extract_invoice_from_pdf(filepath)
    data['file_path'] = filename
    return jsonify(data)


def _export_invoices_csv(db, year):
    rows = db.execute(
        "SELECT * FROM invoices WHERE strftime('%Y', date)=? ORDER BY date DESC",
        (str(year),)
    ).fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Date','Vendor','Description','Category','Amount (inc GST)','GST',
                'PO Number','Invoice #','Status','Location','Notes'])
    for r in rows:
        w.writerow([r['date'], r['vendor'], r['description'], r['category'],
                    r['amount'], r['gst'] or '', r['po_number'] or '',
                    r['invoice_number'] or '', r['payment_status'],
                    r['location'] or '', r['notes'] or ''])
    return Response(out.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=invoices_{year}.csv'})


@app.route('/budget/new', methods=['GET', 'POST'])
def invoice_new():
    if request.method == 'POST':
        _save_invoice(None)
        flash('Invoice added.', 'success')
        return redirect(url_for('budget'))
    return render_template('invoice_form.html', invoice=None, title='Add Invoice')


@app.route('/budget/<int:id>/edit', methods=['GET', 'POST'])
def invoice_edit(id):
    inv = g.db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    if not inv:
        return redirect(url_for('budget'))
    if request.method == 'POST':
        _save_invoice(id)
        flash('Invoice updated.', 'success')
        return redirect(url_for('budget'))
    return render_template('invoice_form.html', invoice=inv, title='Edit Invoice')


@app.route('/budget/<int:id>/delete', methods=['POST'])
def invoice_delete(id):
    g.db.execute("DELETE FROM invoices WHERE id=?", (id,))
    g.db.commit()
    flash('Invoice deleted.', 'info')
    return redirect(url_for('budget'))


@app.route('/budget/uploads/<filename>')
def invoice_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


def _save_invoice(id):
    f  = request.form
    db = g.db
    amount  = f.get('amount') or 0
    gst_raw = f.get('gst') or None

    file_path = None
    if 'invoice_file' in request.files:
        file = request.files['invoice_file']
        if file and file.filename and allowed_file(file.filename):
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{ts}_{secure_filename(file.filename)}"
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            file_path = filename

    if id is not None and file_path is None:
        existing = db.execute("SELECT file_path FROM invoices WHERE id=?", (id,)).fetchone()
        if existing:
            file_path = existing['file_path']

    if file_path is None:
        extracted = f.get('extracted_file_path', '').strip()
        if extracted:
            file_path = extracted

    vals = (
        f.get('date') or date.today().isoformat(),
        f.get('vendor', '').strip(),
        f.get('description', '').strip(),
        f.get('category', 'ICT Equipment'),
        float(amount),
        float(gst_raw) if gst_raw else None,
        f.get('po_number', '').strip(),
        f.get('invoice_number', '').strip(),
        f.get('payment_status', 'paid'),
        f.get('notes', '').strip(),
        file_path,
        f.get('location', '').strip(),
    )
    if id is None:
        db.execute("""INSERT INTO invoices
            (date,vendor,description,category,amount,gst,po_number,invoice_number,
             payment_status,notes,file_path,location)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", vals)
    else:
        db.execute("""UPDATE invoices SET
            date=?,vendor=?,description=?,category=?,amount=?,gst=?,
            po_number=?,invoice_number=?,payment_status=?,notes=?,
            file_path=?,location=? WHERE id=?""", vals + (id,))
    db.commit()


# ─── Furniture ────────────────────────────────────────────────────────────────

@app.route('/furniture')
def furniture():
    db = g.db
    search = request.args.get('q', '').strip()
    cat = request.args.get('cat', '')
    status = request.args.get('status', '')
    location = request.args.get('location', '')

    query = "SELECT * FROM furniture WHERE 1=1"
    params = []
    if search:
        query += " AND (name LIKE ? OR supplier LIKE ? OR location LIKE ?)"
        p = f'%{search}%'
        params.extend([p, p, p])
    if cat:
        query += " AND category=?"
        params.append(cat)
    if status:
        query += " AND status=?"
        params.append(status)
    if location:
        query += " AND location LIKE ?"
        params.append(f'%{location}%')
    query += " ORDER BY category, name"
    rows = db.execute(query, params).fetchall()

    total_value = sum((r['purchase_price'] or 0) * (r['quantity'] or 1) for r in rows)
    total_items = sum(r['quantity'] or 1 for r in rows)
    return render_template('furniture.html', items=rows,
                           total_value=total_value, total_items=total_items,
                           search=search, cat=cat, status=status, location=location)


@app.route('/furniture/new', methods=['GET', 'POST'])
def furniture_new():
    if request.method == 'POST':
        _save_furniture(None)
        flash('Furniture item added.', 'success')
        return redirect(url_for('furniture'))
    return render_template('furniture_form.html', item=None, title='Add Furniture')


@app.route('/furniture/<int:id>/edit', methods=['GET', 'POST'])
def furniture_edit(id):
    item = g.db.execute("SELECT * FROM furniture WHERE id=?", (id,)).fetchone()
    if not item:
        return redirect(url_for('furniture'))
    if request.method == 'POST':
        _save_furniture(id)
        flash('Furniture item updated.', 'success')
        return redirect(url_for('furniture'))
    return render_template('furniture_form.html', item=item, title='Edit Furniture')


@app.route('/furniture/<int:id>/delete', methods=['POST'])
def furniture_delete(id):
    g.db.execute("DELETE FROM furniture WHERE id=?", (id,))
    g.db.commit()
    flash('Item deleted.', 'info')
    return redirect(url_for('furniture'))


def _save_furniture(id):
    f = request.form
    db = g.db
    now = datetime.now().isoformat()
    vals = (
        f.get('name', '').strip(),
        f.get('category', 'Other'),
        f.get('quantity') or 1,
        f.get('location', '').strip(),
        f.get('condition', 'Good'),
        f.get('purchase_date') or None,
        f.get('purchase_price') or None,
        f.get('supplier', '').strip(),
        f.get('status', 'active'),
        f.get('notes', '').strip(),
        now,
    )
    if id is None:
        db.execute("""INSERT INTO furniture
            (name,category,quantity,location,condition,purchase_date,purchase_price,
             supplier,status,notes,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""", vals)
    else:
        db.execute("""UPDATE furniture SET
            name=?,category=?,quantity=?,location=?,condition=?,purchase_date=?,
            purchase_price=?,supplier=?,status=?,notes=?,updated_at=? WHERE id=?""",
            vals + (id,))
    db.commit()


# ─── Auto Backup ──────────────────────────────────────────────────────────────

BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'backups')
BACKUP_INTERVAL_DAYS = 7
MAX_BACKUPS = 4


def _do_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = os.path.join(BACKUP_DIR, f'tracker_{ts}.db')
    shutil.copy2(DB_PATH, dst)

    files = sorted(
        [f for f in os.listdir(BACKUP_DIR) if f.startswith('tracker_') and f.endswith('.db')]
    )
    while len(files) > MAX_BACKUPS:
        os.remove(os.path.join(BACKUP_DIR, files.pop(0)))

    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('last_backup', ?)",
                 (datetime.now().isoformat(),))
    conn.commit()
    conn.close()
    print(f'[backup] Saved {dst}')


def _backup_worker():
    time.sleep(10)
    while True:
        try:
            conn = get_db()
            row = conn.execute("SELECT value FROM settings WHERE key='last_backup'").fetchone()
            conn.close()
            last = datetime.fromisoformat(row['value']) if row else None
            if last is None or (datetime.now() - last).days >= BACKUP_INTERVAL_DAYS:
                _do_backup()
        except Exception as e:
            print(f'[backup] Error: {e}')
        time.sleep(3600)


if __name__ == '__main__':
    init_db()
    t = threading.Thread(target=_backup_worker, daemon=True)
    t.start()
    app.run(debug=True, port=5000, host='0.0.0.0')
